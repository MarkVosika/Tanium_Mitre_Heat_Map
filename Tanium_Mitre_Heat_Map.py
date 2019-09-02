#import needed libraries
import os
import re
import sys
import json
import binascii
from cryptography.fernet import Fernet
import requests 
import urllib3
from requests.auth import HTTPBasicAuth
requests.packages.urllib3.disable_warnings()
from stix2 import TAXIICollectionSource, Filter
from taxii2client import Server, Collection
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

#________________________________________________________________________________________________________________________________________________

#empty variables for handling encryption
key = ""
uncipher_text = ""
cipher_suite = ""
encryptedpwd = ""

#authentication variables
base_url = 'https://<server>.com'
username = '<base64_username>'
key = '<base64_fernet_key>'
cipher_suite = Fernet(key)
with open('<path to cipher text file>', 'rb') as file_object1:
	for line in file_object1:
		encryptedpwd = line

#authenticate to host
http = urllib3.PoolManager()
handshake = HTTPBasicAuth(binascii.a2b_base64(username).strip(), binascii.a2b_base64((cipher_suite.decrypt(encryptedpwd))).strip())
r = requests.post(base_url + '/auth',verify=False,auth=handshake)
sessionid = r.content
file_object1.close()

#Use existing session to get IOC information and load into python
intels = requests.get(base_url + '/plugin/products/detect3/api/v1/intels',verify=False, headers={'session': sessionid})
json_input = (json.dumps(intels.json(), indent=4, sort_keys=True, ensure_ascii=False))
json_load = json.loads(json_input)

#empty lists for storing platform specific IOC names
all_ioc = []
windows_ioc = []
linux_ioc = []
mac_ioc = []

#Find IOC name, search for technique id patttern and platform, if found assign IOC name to platform specific list
for lst in json_load:
	for k,v in lst.items():
		if k == 'name' and re.search('T\d\d\d\d', lst['name'], re.IGNORECASE):
			all_ioc.append(lst['name'])
		if k == 'name' and re.search('T\d\d\d\d', lst['name'], re.IGNORECASE) and re.search('windows', lst['name'], re.IGNORECASE):
			windows_ioc.append(lst['name'])
		if k == 'name' and re.search('T\d\d\d\d', lst['name'], re.IGNORECASE) and re.search('linux', lst['name'], re.IGNORECASE):
			linux_ioc.append(lst['name'])
		if k == 'name' and re.search('T\d\d\d\d', lst['name'], re.IGNORECASE) and re.search('macOS', lst['name'], re.IGNORECASE):
			mac_ioc.append(lst['name'])

#_______________________________________________________________________________________________________________________________________

#proxy awareness (if not needed remove "proxies" from all lines)
proxies = {'http': "http://" + binascii.a2b_base64(username).strip() + ":" + binascii.a2b_base64((cipher_suite.decrypt(encryptedpwd))).strip() + "@<Proxy_Server:Port>", 'https': "http://" + binascii.a2b_base64(username).strip() + ":" + binascii.a2b_base64((cipher_suite.decrypt(encryptedpwd))).strip() + "@<Proxy_Server:Port>"}


#enterprise attack source
collection = Collection("https://cti-taxii.mitre.org/stix/collections/95ecc380-afe9-11e4-9b6c-751b66dd541e/", proxies = proxies)
# supply the TAXII2 collection to TAXIICollection
tc_source = TAXIICollectionSource(collection)

#filter to only techniques
filt = Filter('type', '=', 'attack-pattern')
techniques = tc_source.query([filt])


#generate list of all kill chain phases
kc_list = []

for technique in techniques:
	for k,v in technique.items():
		for i in technique["kill_chain_phases"]:
			for k,v in i.items():
				kc_list.append(v.encode("utf-8"))

#remove duplicates and sort
deduped_kc = sorted(list(dict.fromkeys(kc_list)))
deduped_kc.remove('mitre-attack')

#seperate platforms
windows = []
linux = []
mac = []
all_techniques = []

for technique in techniques:
    for k,v in technique.items():
    	if k == 'x_mitre_platforms':
    		all_techniques.append(technique)
    		if 'Windows' in v:
    			windows.append(technique)			
    		if "Linux" in v:
    			linux.append(technique)
    		if "macOS" in v:
    			mac.append(technique)


#create worksheet tabs for each platform
count = 1

wb = Workbook()
sheet1 = wb.get_sheet_by_name('Sheet')
sheet1.title = 'All_Techniques'
sheet2 = wb.create_sheet('Windows')
sheet3 = wb.create_sheet('Linux')
sheet4 = wb.create_sheet('macOS')

list_of_lists = []

#_______________________________________________________________________________________________________________________________________________________

def parse_json(platform):
	global count
	#create empty list for each kill chain phase:

	list_of_lists = []

	for i in deduped_kc:
		list_name = []
		list_of_lists.append(list_name)


	#This creates a 3 level deep list of lists.  
	#a list for holding all data, a list for each kill chain phase, and a list for each technique

	for lst in platform:
		for technique in lst:
			for i in technique['external_references']:
				for k,v in i.items():
					if k == 'external_id' and v.startswith('T'): #extracts the technique ID
						t_id = v.encode("utf-8")
			for index in enumerate(deduped_kc):
				for i in technique["kill_chain_phases"]:
					for k,v in i.items():
						if index[1] == v:
							list_of_lists[index[0]].append([t_id, technique["name"].encode("utf-8")]) #list containing technique id and name appended
	
	# write to excel file all the data, making first row bold
	bold_font = Font(bold = True)		#set bold font variable
	#wrap_text = Alignment(wrap_text=True) #optional
	redfill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type = 'solid')

	if count == 1:
		current_sheet = sheet1
		ioc_platform = all_ioc
	elif count == 2:
		current_sheet = sheet2
		ioc_platform = windows_ioc
	elif count == 3:
		current_sheet = sheet3
		ioc_platform = linux_ioc
	elif count == 4:
		current_sheet = sheet4
		ioc_platform = mac_ioc

	current_sheet.append(deduped_kc)
	for cell in current_sheet["1:1"]:
		cell.font = bold_font
		cell.alignment = wrap_text
	
	column = '@'

	for lst in list_of_lists:
		column = chr(ord(column) + 1)
		row = 2
		for technique in lst:
			current_sheet[column + str(row)] = technique[1] # [1] = technique name
			#current_sheet[column + str(row)].alignment = wrap_text #optional
			for i in ioc_platform:
				if technique[0] in i: # [0] = technique ID
					current_sheet[column + str(row)].fill = redfill
			row += 1 

	count += 1

	wb.save('Heat_Map.xlsx')							

#_______________________________________________________________________________________________________________________________________________________

# call the main function
parse_json([all_techniques])
parse_json([windows]) 
parse_json([linux])
parse_json([mac])

